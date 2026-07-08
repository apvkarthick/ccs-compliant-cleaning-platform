from io import BytesIO

from openpyxl import Workbook

from api.excel_parser import parse_chemical_register, parse_client_workbook


def _sample_register_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet["B7"] = "Chemical Register"
    sheet["C8"] = "Biniris C/- Springfield Tower"
    sheet["C9"] = "2026-07-01"
    sheet["C10"] = "Matthew King"
    sheet["C12"] = "0422 228 965"

    sheet["A15"] = "Product Code"
    sheet["B15"] = "Product Name"
    sheet["C15"] = "Hazardous"
    sheet["F15"] = "Risk Assessment"
    sheet["K15"] = "SDS Expiry"
    sheet["N15"] = "Selected"

    sheet["A17"] = "Smart Clean Range"
    sheet["A18"] = "ALLPURP5L"
    sheet["B18"] = "All Purpose Sanitiser Soak"
    sheet["C18"] = "No"
    sheet["F18"] = "No"
    sheet["K18"] = "2029-01-01"
    sheet["N18"] = 1

    sheet["A19"] = "LAUNDPOW15KGF"
    sheet["B19"] = "Laundry Powder"
    sheet["C19"] = "Yes"
    sheet["F19"] = "Yes"
    sheet["K19"] = "2028-05-01"
    sheet["N19"] = 1

    sheet["A20"] = "SKIPME"
    sheet["B20"] = "Skipped Product"
    sheet["N20"] = 0

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def test_parse_chemical_register_extracts_contact_products_and_document_urls():
    source_files = [
        "ALLPURP5L_ All Purpose Sanitiser Soak SDS.pdf",
        "LAUNDPOW15KG_ Laundry Powder SDS.pdf",
        "RISK_LAUNDPOW15KGF_ Laundry Powder Risk Assessment.pdf",
    ]

    preview = parse_chemical_register(
        _sample_register_bytes(),
        source_files=source_files,
        public_base_url="https://ccs.example.test",
    )

    assert preview["customer"]["company"] == "Biniris C/- Springfield Tower"
    assert preview["customer"]["contact_name"] == "Matthew King"
    assert preview["customer"]["phone"] == "0422 228 965"
    assert preview["register"]["title"] == "Chemical Register"
    assert preview["register"]["date"] == "2026-07-01"

    products = preview["products"]
    assert [product["code"] for product in products] == ["ALLPURP5L", "LAUNDPOW15KGF"]
    assert products[0]["sds"]["matched"] is True
    assert products[0]["sds"]["url"] == "https://ccs.example.test/api/documents/source/ALLPURP5L_%20All%20Purpose%20Sanitiser%20Soak%20SDS.pdf"
    assert products[0]["risk_assessment"]["matched"] is False
    assert products[1]["sds"]["matched"] is True
    assert products[1]["risk_assessment"]["matched"] is True


def test_parse_client_workbook_extracts_customers_products_and_pdf_links_across_sheets():
    workbook = Workbook()
    customers = workbook.active
    customers.title = "Customer Data"
    customers.append(["Customer", "Contact Name", "Email", "Phone"])
    customers.append(["Springfield Tower", "Matthew King", "matthew@example.com", "0422 228 965"])
    customers.append(["North Site", "Sarah Jones", "sarah@example.com", "0400 000 001"])

    chemicals = workbook.create_sheet("Chemical SDS Links")
    chemicals.append(
        [
            "Product Code",
            "Chemical Name",
            "Selected",
            "SDS PDF Link",
            "Risk Assessment Link",
            "Chemical Register URL",
        ]
    )
    chemicals.append(
        [
            "ALLPURP5L",
            "All Purpose Sanitiser Soak",
            "yes",
            "https://cdn.example.com/sds/allpurp.pdf",
            "",
            "https://cdn.example.com/register/springfield.pdf",
        ]
    )
    chemicals.append(
        [
            "BLEACH5L",
            "Bleach 4%",
            "yes",
            "https://cdn.example.com/sds/bleach.pdf",
            "https://cdn.example.com/risk/bleach.pdf",
            "",
        ]
    )

    stream = BytesIO()
    workbook.save(stream)

    preview = parse_client_workbook(stream.getvalue(), public_base_url="https://ccs.example.test")

    assert preview["workbook_type"] == "client_workbook"
    assert preview["customer"]["company"] == "Springfield Tower"
    assert preview["customer"]["contact_name"] == "Matthew King"
    assert preview["contacts"] == [
        {"company": "Springfield Tower", "name": "Matthew King", "email": "matthew@example.com", "phone": "0422 228 965"},
        {"company": "North Site", "name": "Sarah Jones", "email": "sarah@example.com", "phone": "0400 000 001"},
    ]
    assert preview["register"]["url"] == "https://cdn.example.com/register/springfield.pdf"
    assert [product["code"] for product in preview["products"]] == ["ALLPURP5L", "BLEACH5L"]
    assert preview["products"][0]["sds"]["url"] == "https://cdn.example.com/sds/allpurp.pdf"
    assert preview["products"][0]["risk_assessment"]["matched"] is False
    assert preview["products"][1]["risk_assessment"]["url"] == "https://cdn.example.com/risk/bleach.pdf"


def test_parse_client_workbook_attaches_site_emails_to_products():
    workbook = Workbook()
    site_a = workbook.active
    site_a.title = "Site A"
    site_a.append(["Customer", "Contact Name", "Email"])
    site_a.append(["Site A Pty Ltd", "Alice A", "alice@example.com"])
    site_a.append([])
    site_a.append(["Product Code", "Chemical Name", "Selected", "SDS PDF Link"])
    site_a.append(["ALLPURP5L", "All Purpose Sanitiser Soak", "yes", "https://cdn.example.com/sds/allpurp.pdf"])

    site_b = workbook.create_sheet("Site B")
    site_b.append(["Customer", "Contact Name", "Email"])
    site_b.append(["Site B Pty Ltd", "Bob B", "bob@example.com"])
    site_b.append([])
    site_b.append(["Product Code", "Chemical Name", "Selected", "SDS PDF Link"])
    site_b.append(["ALLPURP5L", "All Purpose Sanitiser Soak", "yes", "https://cdn.example.com/sds/allpurp.pdf"])

    stream = BytesIO()
    workbook.save(stream)

    preview = parse_client_workbook(stream.getvalue(), public_base_url="https://ccs.example.test")
    products = preview["products"]

    assert len(products) == 2
    assert products[0]["site_emails"] == ["alice@example.com"]
    assert products[1]["site_emails"] == ["bob@example.com"]


def test_parse_client_workbook_risk_document_is_matched_by_risk_code_filename():
    workbook = Workbook()
    chemicals = workbook.active
    chemicals.title = "Chemical SDS Links"
    chemicals.append(
        [
            "Product Code",
            "Chemical Name",
            "Selected",
            "SDS PDF Link",
            "Risk Assessment Link",
        ]
    )
    chemicals.append(
        [
            "BLEACH5L",
            "Bleach 4%",
            "yes",
            "https://cdn.example.com/sds/bleach.pdf",
            "https://cdn.example.com/risk/wrong-risk-link.pdf",
        ]
    )

    stream = BytesIO()
    workbook.save(stream)

    preview = parse_client_workbook(
        stream.getvalue(),
        source_files=["RISK_BLEACH5L_ Bleach 4% Risk Assessment.pdf"],
        public_base_url="https://ccs.example.test",
    )

    risk = preview["products"][0]["risk_assessment"]
    assert risk["matched"] is True
    assert risk["url"] == "https://ccs.example.test/api/documents/source/RISK_BLEACH5L_%20Bleach%204%25%20Risk%20Assessment.pdf"


def test_parse_client_workbook_uses_hyperlink_targets_for_hosted_pdf_urls():
    workbook = Workbook()
    chemicals = workbook.active
    chemicals.title = "Chemical SDS Links"
    chemicals.append(["Product Code", "Chemical Name", "Selected", "SDS PDF Link", "Risk Assessment Link"])
    chemicals.append(["BLEACH5L", "Bleach 4%", "yes", "SDS", "RISK"])

    chemicals["D2"].hyperlink = "https://client.example.com/files/BLEACH5L_bleach_sds.pdf"
    chemicals["E2"].hyperlink = "https://client.example.com/files/RISK_BLEACH5L_bleach_risk.pdf"

    stream = BytesIO()
    workbook.save(stream)

    preview = parse_client_workbook(stream.getvalue(), source_files=[], public_base_url="https://ccs.example.test")
    product = preview["products"][0]

    assert product["sds"]["matched"] is True
    assert product["sds"]["url"] == "https://client.example.com/files/BLEACH5L_bleach_sds.pdf"
    assert product["risk_assessment"]["matched"] is True
    assert product["risk_assessment"]["url"] == "https://client.example.com/files/RISK_BLEACH5L_bleach_risk.pdf"
