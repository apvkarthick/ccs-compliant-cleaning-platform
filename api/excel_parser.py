from __future__ import annotations

import re
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any
from urllib.parse import quote

from openpyxl import load_workbook


PRODUCT_START_ROW = 16
PRODUCT_END_ROW = 120


def parse_chemical_register(
    workbook_bytes: bytes,
    *,
    source_files: list[str] | None = None,
    public_base_url: str = "",
) -> dict[str, Any]:
    preview = _parse_chemical_register_layout(
        workbook_bytes,
        source_files=source_files,
        public_base_url=public_base_url,
    )
    preview["workbook_type"] = "chemical_register"
    preview.setdefault("contacts", _contacts_from_customer(preview["customer"]))
    return preview


def parse_client_workbook(
    workbook_bytes: bytes,
    *,
    source_files: list[str] | None = None,
    public_base_url: str = "",
) -> dict[str, Any]:
    structured_preview = _parse_structured_workbook(
        workbook_bytes,
        source_files=source_files,
        public_base_url=public_base_url,
    )
    if structured_preview["products"]:
        return structured_preview
    return parse_chemical_register(
        workbook_bytes,
        source_files=source_files,
        public_base_url=public_base_url,
    )


def _parse_chemical_register_layout(
    workbook_bytes: bytes,
    *,
    source_files: list[str] | None = None,
    public_base_url: str = "",
) -> dict[str, Any]:
    workbook = load_workbook(BytesIO(workbook_bytes), data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    files = source_files or []

    products = []
    for row_number in range(PRODUCT_START_ROW, PRODUCT_END_ROW + 1):
        code = _clean(sheet.cell(row=row_number, column=1).value)
        name = _clean(sheet.cell(row=row_number, column=2).value)
        if not code or not name:
            continue
        if _looks_like_header_or_category(code, name):
            continue
        if not _is_selected(sheet.cell(row=row_number, column=14).value):
            continue

        sds_file = _match_sds_file(code, files)
        risk_file = _match_risk_file(code, files)

        products.append(
            {
                "row": row_number,
                "code": code,
                "name": name,
                "hazardous": _clean(sheet.cell(row=row_number, column=3).value),
                "un_number": _clean(sheet.cell(row=row_number, column=4).value),
                "max_quantity": _clean(sheet.cell(row=row_number, column=5).value),
                "risk_required": _truthy_label(sheet.cell(row=row_number, column=6).value),
                "hazchem": _clean(sheet.cell(row=row_number, column=7).value),
                "class": _clean(sheet.cell(row=row_number, column=8).value),
                "packing_group": _clean(sheet.cell(row=row_number, column=9).value),
                "use": _clean(sheet.cell(row=row_number, column=10).value),
                "sds_expiry": _clean(sheet.cell(row=row_number, column=11).value),
                "sds": _document_result(sds_file, public_base_url),
                "risk_assessment": _document_result(risk_file, public_base_url),
            }
        )

    missing_documents = [
        {
            "code": product["code"],
            "name": product["name"],
            "missing": [
                label
                for label, key in [("SDS", "sds"), ("Risk Assessment", "risk_assessment")]
                if not product[key]["matched"] and (key == "sds" or product["risk_required"])
            ],
        }
        for product in products
    ]

    return {
        "customer": {
            "company": _clean(sheet["C8"].value),
            "contact_name": _clean(sheet["C10"].value),
            "phone": _clean(sheet["C12"].value),
            "email": _find_email(sheet),
        },
        "register": {
            "title": _clean(sheet["B7"].value) or "Chemical Register",
            "date": _clean(sheet["C9"].value),
            "sheet": sheet.title,
            "product_count": len(products),
            "url": "",
        },
        "products": products,
        "missing_documents": [item for item in missing_documents if item["missing"]],
    }


def _parse_structured_workbook(
    workbook_bytes: bytes,
    *,
    source_files: list[str] | None,
    public_base_url: str,
) -> dict[str, Any]:
    workbook = load_workbook(BytesIO(workbook_bytes), data_only=True)
    files = source_files or []
    contacts: list[dict[str, str]] = []
    products: list[dict[str, Any]] = []
    sheet_contacts: dict[str, set[str]] = {}
    cust_num_to_email: dict[str, str] = {}
    register_url = ""

    for sheet in workbook.worksheets:
        # Try transposed layout (field labels in col A, customer data in cols B+)
        transposed, num_map = _parse_transposed_customer_sheet(sheet)
        if transposed:
            contacts.extend(transposed)
            cust_num_to_email.update(num_map)

        for header_row, headers in _header_rows(sheet):
            if _is_customer_header(headers):
                parsed_contacts = _contacts_from_sheet(sheet, header_row, headers)
                contacts.extend(parsed_contacts)
                emails = {c.get("email", "").strip().lower() for c in parsed_contacts if c.get("email")}
                if emails:
                    sheet_contacts.setdefault(sheet.title, set()).update(emails)
            if _is_product_header(headers):
                parsed_products, parsed_register_url = _products_from_sheet(
                    sheet,
                    header_row,
                    headers,
                    source_files=files,
                    public_base_url=public_base_url,
                )
                products.extend(parsed_products)
                register_url = register_url or parsed_register_url

    for product in products:
        sheet_name = str(product.get("sheet", ""))
        # Match "Cust 38 - ..." sheet names to the customer number → email map
        cust_match = re.search(r"cust(?:omer)?\s*[\-#]?\s*(\d+)", sheet_name, re.IGNORECASE)
        if cust_match and cust_num_to_email:
            email = cust_num_to_email.get(cust_match.group(1), "")
            product["site_emails"] = [email] if email else []
        else:
            product["site_emails"] = sorted(sheet_contacts.get(sheet_name, set()))

    customer = _customer_from_contacts(contacts)
    return {
        "workbook_type": "client_workbook",
        "customer": customer,
        "contacts": contacts,
        "register": {
            "title": "Client Workbook",
            "date": "",
            "sheet": ", ".join(workbook.sheetnames),
            "product_count": len(products),
            "url": register_url,
        },
        "products": products,
        "missing_documents": _missing_documents(products),
    }


def list_source_documents(source_dir: Path) -> list[str]:
    if not source_dir.exists():
        return []
    return sorted(path.name for path in source_dir.iterdir() if path.is_file())


def _clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _looks_like_header_or_category(code: str, name: str) -> bool:
    lowered = f"{code} {name}".lower()
    return "product code" in lowered or code.lower().endswith("range")


def _is_selected(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value == 1
    return str(value).strip().lower() in {"1", "yes", "y", "true", "selected", "x"}


def _truthy_label(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value > 0
    return str(value).strip().lower() in {"yes", "y", "true", "required", "available", "1"}


def _document_result(filename: str | None, public_base_url: str) -> dict[str, Any]:
    if not filename:
        return {"matched": False, "filename": None, "url": None}
    base = public_base_url.rstrip("/")
    quoted = quote(filename)
    return {
        "matched": True,
        "filename": filename,
        "url": f"{base}/api/documents/source/{quoted}" if base else f"/api/documents/source/{quoted}",
    }


def _url_document_result(url: str) -> dict[str, Any]:
    if not url:
        return {"matched": False, "filename": None, "url": None}
    return {"matched": True, "filename": url.rsplit("/", 1)[-1] or url, "url": url}


def _match_sds_file(code: str, files: list[str]) -> str | None:
    aliases = _code_aliases(code)
    return _first_matching_file(files, aliases, risk=False)


def _match_risk_file(code: str, files: list[str]) -> str | None:
    aliases = [f"risk_{alias}" for alias in _code_aliases(code)]
    return _first_matching_file(files, aliases, risk=True)


def _code_aliases(code: str) -> list[str]:
    normalized = _normalize(code)
    aliases = [normalized]
    if normalized.endswith("f"):
        aliases.append(normalized[:-1])
    return aliases


def _first_matching_file(files: list[str], aliases: list[str], *, risk: bool) -> str | None:
    for filename in files:
        normalized = _normalize(filename)
        is_risk_file = normalized.startswith("risk_")
        if risk != is_risk_file:
            continue
        if any(normalized.startswith(alias) for alias in aliases):
            return filename
    return None


def _normalize(value: str) -> str:
    return "".join(char.lower() if char.isalnum() else "_" for char in value).strip("_")


def _find_email(sheet: Any) -> str:
    for row in sheet.iter_rows(min_row=1, max_row=25, values_only=True):
        for value in row:
            text = _clean(value)
            if _looks_like_email(text):
                return text
    return ""


def _looks_like_url(value: str) -> bool:
    return bool(value) and value.startswith(("http://", "https://", "/"))


def _looks_like_email(value: str) -> bool:
    text = _clean(value)
    return "@" in text and "." in text


def _header_rows(sheet: Any) -> list[tuple[int, dict[str, int]]]:
    header_rows = []
    for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 30)):
        headers = {
            _normalize_header(cell.value): cell.column
            for cell in row
            if _normalize_header(cell.value)
        }
        if len(headers) >= 2:
            header_rows.append((row[0].row, headers))
    return header_rows


def _is_customer_header(headers: dict[str, int]) -> bool:
    return bool(_find_header(headers, CUSTOMER_EMAIL_HEADERS)) and bool(
        _find_header(headers, CUSTOMER_COMPANY_HEADERS + CUSTOMER_NAME_HEADERS)
    )


def _is_product_header(headers: dict[str, int]) -> bool:
    has_product = bool(_find_header(headers, PRODUCT_CODE_HEADERS) or _find_header(headers, PRODUCT_NAME_HEADERS))
    has_document = bool(_find_header(headers, SDS_URL_HEADERS + RISK_URL_HEADERS + REGISTER_URL_HEADERS))
    return has_product and has_document


def _contacts_from_sheet(sheet: Any, header_row: int, headers: dict[str, int]) -> list[dict[str, str]]:
    contacts = []
    for row_number in range(header_row + 1, sheet.max_row + 1):
        email = _cell_by_headers(sheet, row_number, headers, CUSTOMER_EMAIL_HEADERS)
        company = _cell_by_headers(sheet, row_number, headers, CUSTOMER_COMPANY_HEADERS)
        name = _cell_by_headers(sheet, row_number, headers, CUSTOMER_NAME_HEADERS)
        phone = _cell_by_headers(sheet, row_number, headers, CUSTOMER_PHONE_HEADERS)
        if not any([email, company, name, phone]):
            continue
        if not email or not _looks_like_email(email):
            continue
        contacts.append({"company": company, "name": name, "email": email, "phone": phone})
    return contacts


def _products_from_sheet(
    sheet: Any,
    header_row: int,
    headers: dict[str, int],
    *,
    source_files: list[str],
    public_base_url: str,
) -> tuple[list[dict[str, Any]], str]:
    products = []
    register_url = ""
    for row_number in range(header_row + 1, sheet.max_row + 1):
        code = _cell_by_headers(sheet, row_number, headers, PRODUCT_CODE_HEADERS)
        name = _cell_by_headers(sheet, row_number, headers, PRODUCT_NAME_HEADERS)
        if not code and not name:
            continue

        selected = _cell_by_headers(sheet, row_number, headers, SELECTED_HEADERS)
        if selected and not _is_selected(selected):
            continue

        sds_url = _url_by_headers(sheet, row_number, headers, SDS_URL_HEADERS)
        risk_raw = _url_by_headers(sheet, row_number, headers, RISK_URL_HEADERS)
        row_register_url = _url_by_headers(sheet, row_number, headers, REGISTER_URL_HEADERS)
        register_url = register_url or row_register_url

        # "risk assessment" column may hold a URL or a yes/no flag
        risk_url = risk_raw if _looks_like_url(risk_raw) else ""
        risk_required_by_flag = _truthy_label(risk_raw) if not risk_url else False

        sds_file = _match_sds_file(code, source_files) if code else None
        risk_file = _match_risk_file(code, source_files) if code else None
        sds_url_match = _url_document_result(sds_url) if sds_url else _url_document_result("")
        risk_url_match = _url_document_result(risk_url) if risk_url else _url_document_result("")

        products.append(
            {
                "row": row_number,
                "sheet": sheet.title,
                "code": code,
                "name": name or code,
                "hazardous": _cell_by_headers(sheet, row_number, headers, HAZARD_HEADERS),
                "un_number": _cell_by_headers(sheet, row_number, headers, UN_HEADERS),
                "max_quantity": _cell_by_headers(sheet, row_number, headers, QUANTITY_HEADERS),
                "risk_required": bool(risk_url or risk_file or risk_required_by_flag),
                "hazchem": "",
                "class": "",
                "packing_group": "",
                "use": _cell_by_headers(sheet, row_number, headers, USE_HEADERS),
                "sds_expiry": _cell_by_headers(sheet, row_number, headers, EXPIRY_HEADERS),
                "sds": (
                    sds_url_match
                    if sds_url_match["matched"]
                    else _document_result(sds_file, public_base_url)
                ),
                "risk_assessment": (
                    _document_result(risk_file, public_base_url)
                    if risk_file
                    else (risk_url_match if risk_url_match["matched"] else _document_result(None, public_base_url))
                ),
            }
        )
    return products, register_url


def _cell_by_headers(sheet: Any, row_number: int, headers: dict[str, int], candidates: list[str]) -> str:
    header = _find_header(headers, candidates)
    if not header:
        return ""
    return _clean(sheet.cell(row=row_number, column=headers[header]).value)


def _url_by_headers(sheet: Any, row_number: int, headers: dict[str, int], candidates: list[str]) -> str:
    header = _find_header(headers, candidates)
    if not header:
        return ""
    cell = sheet.cell(row=row_number, column=headers[header])
    if cell.hyperlink and cell.hyperlink.target:
        return _clean(cell.hyperlink.target)
    return _clean(cell.value)


def _find_header(headers: dict[str, int], candidates: list[str]) -> str:
    for candidate in candidates:
        normalized = _normalize_header(candidate)
        if normalized in headers:
            return normalized
    return ""


def _normalize_header(value: Any) -> str:
    return _normalize(_clean(value))


def _customer_from_contacts(contacts: list[dict[str, str]]) -> dict[str, str]:
    if not contacts:
        return {"company": "", "contact_name": "", "phone": "", "email": ""}
    first = contacts[0]
    return {
        "company": first.get("company", ""),
        "contact_name": first.get("name", ""),
        "phone": first.get("phone", ""),
        "email": first.get("email", ""),
    }


def _contacts_from_customer(customer: dict[str, str]) -> list[dict[str, str]]:
    if not customer.get("email"):
        return []
    return [
        {
            "company": customer.get("company", ""),
            "name": customer.get("contact_name", ""),
            "email": customer.get("email", ""),
            "phone": customer.get("phone", ""),
        }
    ]


def _missing_documents(products: list[dict[str, Any]]) -> list[dict[str, Any]]:
    seen: set[tuple[str, str]] = set()
    result = []
    for product in products:
        missing = [
            label
            for label, key in [("SDS", "sds"), ("Risk Assessment", "risk_assessment")]
            if not product[key]["matched"] and (key == "sds" or product["risk_required"])
        ]
        if not missing:
            continue
        dedup_key = (product.get("code") or "", product.get("name") or "")
        if dedup_key in seen:
            continue
        seen.add(dedup_key)
        result.append({"code": product["code"], "name": product["name"], "missing": missing})
    return result


def _parse_transposed_customer_sheet(sheet: Any) -> tuple[list[dict[str, str]], dict[str, str]]:
    """
    Parse sheets where field labels are in column A and each column B+ is a customer.
    Returns (contacts, customer_number_to_email_map).
    Example layout:
        A1: Customer No. | B1: 38  | C1: 248 | D1: 391
        A3: Site Name    | B3: ... | C3: ... | D3: ...
        A5: Email        | B5: a@b | C5: c@d | D5: e@f
    """
    _email_keys = {_normalize_header(h) for h in ["email", "email address", "contact email", "customer email"]}
    _company_keys = {_normalize_header(h) for h in ["site name", "company", "company name", "customer name", "site", "client"]}
    _name_keys = {_normalize_header(h) for h in ["contact", "contact name", "name"]}
    _phone_keys = {_normalize_header(h) for h in ["phone", "phone number", "mobile"]}
    _cust_num_keys = {"customer_no", "cust_no", "customer_number", "customer_num", "customer_no_"}

    email_row = company_row = name_row = phone_row = cust_num_row = None

    for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 30)):
        label = _normalize_header(row[0].value)
        if not label:
            continue
        if label in _email_keys:
            email_row = row
        elif label in _company_keys:
            company_row = row
        elif label in _name_keys:
            name_row = row
        elif label in _phone_keys:
            phone_row = row
        elif label in _cust_num_keys or label.startswith("customer_n"):
            cust_num_row = row

    if not email_row:
        return [], {}

    contacts: list[dict[str, str]] = []
    cust_num_to_email: dict[str, str] = {}

    for col_idx in range(1, len(email_row)):
        email = _clean(email_row[col_idx].value)
        if not email or not _looks_like_email(email):
            continue
        company = _clean(company_row[col_idx].value) if company_row is not None else ""
        name = _clean(name_row[col_idx].value) if name_row is not None else ""
        phone = _clean(phone_row[col_idx].value) if phone_row is not None else ""
        contacts.append({"company": company, "name": name, "email": email, "phone": phone})
        if cust_num_row is not None:
            num = _clean(cust_num_row[col_idx].value)
            if num:
                cust_num_to_email[num] = email

    return contacts, cust_num_to_email


CUSTOMER_COMPANY_HEADERS = ["customer", "customer name", "company", "company name", "client", "site", "site name"]
CUSTOMER_NAME_HEADERS = ["contact", "contact name", "name", "customer contact", "primary contact"]
CUSTOMER_EMAIL_HEADERS = ["email", "email address", "contact email", "customer email"]
CUSTOMER_PHONE_HEADERS = ["phone", "phone number", "mobile", "contact phone"]

PRODUCT_CODE_HEADERS = ["product code", "code", "sku", "item code", "chemical code"]
PRODUCT_NAME_HEADERS = ["product", "product name", "chemical", "chemical name", "sds product", "name of substance", "substance name"]
SDS_URL_HEADERS = ["sds url", "sds link", "sds pdf link", "sds pdf url", "sds pdf", "sds"]
RISK_URL_HEADERS = ["risk assessment url", "risk assessment link", "risk url", "risk link", "risk assessment"]
REGISTER_URL_HEADERS = ["chemical register url", "chemical register link", "register url", "register link"]
SELECTED_HEADERS = ["selected", "include", "send", "active", "enabled"]
HAZARD_HEADERS = ["hazardous", "hazard", "dangerous good"]
UN_HEADERS = ["un", "un number", "un no"]
QUANTITY_HEADERS = ["max quantity", "quantity", "maximum quantity"]
USE_HEADERS = ["use", "usage", "application"]
EXPIRY_HEADERS = ["sds expiry", "sds expiry date", "expiry", "expiry date"]
