"""Site-level SDS/Risk document distribution based on client-supplied mapping files."""
from __future__ import annotations

import io
import json
import os
import re
from datetime import datetime, timezone
from typing import Any
from urllib.error import HTTPError
from urllib.parse import quote
from urllib.request import Request, urlopen

import pandas as pd

from .distribution import (
    _render_branded_html,
    email_open_pixel_url,
    tracking_url,
)

_BATCH = 500


# ---------------------------------------------------------------------------
# Excel parsers
# ---------------------------------------------------------------------------

def _first_val(row: Any, *candidates: str) -> str:
    """Return first non-empty value from row matching any candidate column name (case-insensitive)."""
    cols = {str(k).upper().strip(): v for k, v in row.items()}
    for c in candidates:
        v = str(cols.get(c.upper(), "")).strip()
        if v and v != "nan":
            return v
    return ""


def parse_mapping_excel(data: bytes) -> list[dict[str, Any]]:
    df = pd.read_excel(io.BytesIO(data), dtype=str)
    # Normalise column headers to uppercase so mismatched casing never silently drops data
    df.columns = [str(c).upper().strip() for c in df.columns]
    sites: list[dict[str, Any]] = []
    for _, row in df.iterrows():
        accno = _first_val(row, "ACCNO", "ACCOUNT_NO", "ACCOUNT", "ACC_NO", "ACCT")
        if not accno:
            continue
        name = _first_val(row, "NAME", "CUST_NAME", "CUSTOMER_NAME", "CUSTOMER",
                          "SITE_NAME", "SITENAME", "ACCT_NAME", "ACCOUNT_NAME", "COMPANY")
        ho_name = _first_val(row, "HO_NAME", "HEAD_OFFICE", "HEAD_OFFICE_NAME",
                             "HEADOFFICE", "PARENT_NAME", "PARENT", "HO")
        ho_accno = _first_val(row, "HO_ACCNO", "HEAD_OFFICE_ACCNO", "HO_ACCOUNT", "PARENT_ACCNO")
        raw_email = _first_val(row, "EMAIL", "CONT_EMAIL", "CONTACT_EMAIL",
                               "EMAIL_ADDRESS", "EMAILS", "CONT_EMAILS")
        raw_codes = _first_val(row, "STOCKCODES", "STOCK_CODES", "PRODUCTS",
                               "PRODUCT_CODES", "CODES", "ITEMS")
        emails = [e.strip() for e in raw_email.split(";") if e.strip()]
        stockcodes = [s.strip() for s in raw_codes.split(",") if s.strip()]
        sites.append({
            "accno": accno,
            "ho_accno": ho_accno,
            "ho_name": ho_name,
            "name": name,
            "emails": emails,
            "stockcodes": stockcodes,
        })
    return sites


def parse_sds_links(sds_data: bytes | None = None, risk_data: bytes | None = None) -> list[dict[str, Any]]:
    """Parse SDS URL file + Risk URL file → merged list of {stock_code, sds_url, risk_url}.
    Either argument may be None or empty — only non-empty files are parsed."""
    sds_map: dict[str, str] = {}
    risk_map: dict[str, str] = {}

    if sds_data:
        df_sds = pd.read_excel(io.BytesIO(sds_data), header=None, dtype=str)
        for url in df_sds[0]:
            url = str(url).strip()
            if not url or url == "nan":
                continue
            fname = url.split("/")[-1]
            m = re.match(r"([A-Z0-9]+)_", fname)
            if m:
                sds_map[m.group(1)] = url

    if risk_data:
        df_risk = pd.read_excel(io.BytesIO(risk_data), header=None, dtype=str)
        for url in df_risk[0]:
            url = str(url).strip()
            if not url or url == "nan":
                continue
            fname = url.split("/")[-1]
            m = re.match(r"RISK_([A-Z0-9]+)_", fname)
            if m:
                risk_map[m.group(1)] = url

    all_codes = sorted(set(sds_map) | set(risk_map))
    return [
        {
            "stock_code": code,
            "sds_url": sds_map.get(code),
            "risk_url": risk_map.get(code),
        }
        for code in all_codes
    ]


def parse_chemical_register(data: bytes) -> list[dict[str, Any]]:
    """Parse Chemical Register xlsx → full product records including metadata columns.

    Handles two layouts:
    - Standard: "PRODUCT CODE" is a header cell; product code is in that column.
    - Title Sheet: product code is in col A (no header label); "NAME OF SUBSTANCE"
      identifies the header row. Section-divider rows (containing spaces) are skipped.
    """
    df = pd.read_excel(io.BytesIO(data), header=None, dtype=str)
    header_row = None
    for i, row in df.iterrows():
        row_upper = [str(v).strip().upper() for v in row]
        if any(v == "PRODUCT CODE" or "NAME OF SUBSTANCE" in v for v in row_upper if v):
            header_row = i
            break
    if header_row is None:
        raise ValueError("Chemical Register: header row not found (expected 'PRODUCT CODE' or 'NAME OF SUBSTANCE')")

    # Normalize headers — uppercase + collapse internal whitespace; rename blank/nan headers
    # to avoid duplicate column names (pandas Series.get on a duplicate returns a Series, not scalar).
    raw_cols = [" ".join(str(v).strip().upper().split()) for v in df.iloc[header_row]]
    seen: dict[str, int] = {}
    clean_cols: list[str] = []
    for c in raw_cols:
        if c in ("NAN", "", "NONE"):
            n = seen.get("__UNNAMED__", 0)
            clean_cols.append(f"__UNNAMED_{n}__")
            seen["__UNNAMED__"] = n + 1
        else:
            clean_cols.append(c)
    df.columns = clean_cols
    df = df.iloc[header_row + 1:].reset_index(drop=True)

    def _col(*candidates: str) -> str | None:
        for c in candidates:
            if c in df.columns:
                return c
        return None

    def _col_contains(substring: str) -> str | None:
        sub = substring.upper()
        return next((c for c in df.columns if sub in c), None)

    # Standard format: "PRODUCT CODE" header; Title Sheet: col A is blank → renamed __UNNAMED_0__
    code_col = _col("PRODUCT CODE") or (clean_cols[0] if clean_cols else df.columns[0])

    name_col = (
        _col(
            "PRODUCT NAME", "CHEMICAL NAME", "PRODUCT / CHEMICAL NAME", "PRODUCT/CHEMICAL NAME",
            "NAME OF SUBSTANCE", "SUBSTANCE NAME",
        )
        or _col_contains("CHEMICAL / PRODUCT")  # master register: "CHEMICAL / PRODUCT - SmartClean Range"
    )
    hazard_col = (
        _col(
            "HAZARD CLASSIFICATION", "HAZARD CLASS", "GHS CLASSIFICATION", "CLASSIFICATION",
            "HAZARD STATUS",
            "IS IT CLASSED AS HAZARDOUS OR DANGEROUS? (Y/N & H/D)",
            "IS IT CLASSED AS HAZARDOUS OR DANGEROUS?",
        )
        or _col_contains("HAZARDOUS")
    )
    use_col = _col(
        "PRIMARY USE", "APPLICATION", "USE", "PRIMARY USE / APPLICATION", "PRIMARY USE/APPLICATION",
        "WHAT IS IT USED FOR",
    )
    signal_col = _col("SIGNAL WORD", "SIGNAL")
    un_col = _col("UN NO", "UN NUMBER", "UN #", "UN#", "UN NO.")
    risk_col = (
        _col(
            "RISK ASSESSMENT", "RISK ASSESS", "RISK ASSESSMENT REQUIRED",
            "RISK ASSESSMENT Y/N", "RISK ASSESSMENT (Y/N)", "RA REQUIRED",
            "REQUIRES RISK ASSESSMENT", "RISK ASSESSMENT REQUIRED?", "RA",
            "RISK ASSESSMENT AVAILABLE (YES/NO) LOCATED BEHIND MSDS",
            "RISK ASSESSMENT AVAILABLE",
        )
        or _col_contains("RISK ASSESSMENT")
    )
    expiry_col = _col(
        "SDS REVIEW DATE", "SDS EXPIRY", "SDS REVIEW", "REVIEW DATE", "EXPIRY DATE",
        "SDS EXPIRY DATE",
    )
    maxqty_col = _col("MAXIMUM QTY", "MAXIMUM QUANTITY", "MAX QTY", "MAX. QTY")
    hazchem_col = _col("HAZCHEM", "HAZCHEM CODE")
    class_col = _col("CLASS", "DG CLASS", "HAZARD CLASS CODE")
    packgroup_col = _col("PACKING GROUP", "PACK GROUP", "PG")

    records: list[dict[str, Any]] = []
    for _, row in df.iterrows():
        code = str(row.get(code_col, "")).strip()
        if not code or code == "nan":
            continue
        # Skip section-divider rows (e.g. "Smart Clean Range") — product codes have no spaces
        if " " in code:
            continue

        risk_val = str(row.get(risk_col, "") if risk_col else "").strip().upper()
        risk_required = risk_val in ("YES", "Y", "TRUE", "1", "X")

        raw_expiry = str(row.get(expiry_col, "") if expiry_col else "").strip()
        sds_expiry = None
        if raw_expiry and raw_expiry != "nan":
            for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"):
                try:
                    from datetime import datetime as _dt
                    sds_expiry = _dt.strptime(raw_expiry.split(" ")[0], fmt).date().isoformat()
                    break
                except ValueError:
                    continue

        def _val(col: str | None) -> str | None:
            if not col:
                return None
            v = str(row.get(col, "")).strip()
            return v if v and v != "nan" else None

        records.append({
            "stock_code": code,
            "risk_assessment_required": risk_required,
            "sds_expiry": sds_expiry,
            "product_name": _val(name_col),
            "hazard_classification": _val(hazard_col),
            "primary_use": _val(use_col),
            "signal_word": _val(signal_col),
            "un_number": _val(un_col),
            "maximum_qty": _val(maxqty_col),
            "hazchem": _val(hazchem_col),
            "chemical_class": _val(class_col),
            "packing_group": _val(packgroup_col),
        })
    return records


def fetch_product_metadata(stock_codes: list[str]) -> dict[str, dict]:
    """Return product metadata keyed by stock_code.

    For codes with no direct metadata, falls back to any other code in the same
    stock-group row (col A, B, C, D... in the size-mapping sheet) that does have
    metadata.  The customer's own code is preserved as the dict key so the
    Chemical Register always shows the customer's actual stock code.
    """
    if not stock_codes:
        return {}

    _FIELDS = (
        "stock_code,product_name,hazard_classification,primary_use,"
        "signal_word,un_number,risk_assessment_required,sds_expiry,"
        "maximum_qty,hazchem,chemical_class,packing_group"
    )

    def _in_filter(codes: list[str] | set[str]) -> str:
        """Build a PostgREST in.() value list.
        Percent-encode special URL chars (&, comma, etc.) in each value so the
        query string stays valid; the HTTP layer decodes them before PostgREST
        parses the filter.  Avoids double-quoted syntax which older PostgREST
        versions do not support."""
        from urllib.parse import quote as _q
        safe = ",".join(_q(str(c), safe="") for c in codes)
        return f"in.({safe})"

    # Direct lookup
    rows = _sb_get("ccs_sds_links", f"select={_FIELDS}&stock_code={_in_filter(stock_codes)}")
    result: dict[str, dict] = {r["stock_code"]: r for r in rows if r.get("product_name")}

    missing = [c for c in stock_codes if c not in result]
    if not missing:
        return result

    # Build group membership map: code → all codes in the same group row
    groups = _sb_get("ccs_stock_groups", "select=primary_code,related_codes")
    group_members: dict[str, list[str]] = {}
    for g in groups:
        members = [g["primary_code"]] + (g.get("related_codes") or [])
        for code in members:
            group_members[code] = members

    # Collect all alternative codes to fetch in one query
    alt_needed: set[str] = set()
    for code in missing:
        for alt in group_members.get(code, []):
            if alt != code:
                alt_needed.add(alt)

    if alt_needed:
        alt_rows = _sb_get("ccs_sds_links", f"select={_FIELDS}&stock_code={_in_filter(alt_needed)}")
        alt_meta: dict[str, dict] = {r["stock_code"]: r for r in alt_rows if r.get("product_name")}

        for code in missing:
            for alt in group_members.get(code, []):
                if alt in alt_meta:
                    result[code] = alt_meta[alt]  # use alt's metadata, but key stays as customer code
                    break

    return result


_PREPARED_BY = "Matthew King Compliant Cleaning Supplies & Systems PTY LTD  Ph 1300 314 491"


def generate_chemical_register_excel(
    site_name: str,
    accno: str,
    stock_codes: list[str],
    today: str,
    register_codes: set[str] | None = None,
) -> bytes:
    """Generate a per-site Chemical Register Excel (columns A–K) with branded cover page.

    Only includes codes present in the Chemical Register (ccs_sds_links). Output is
    sorted alphabetically by product code.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    metadata = fetch_product_metadata(stock_codes)

    # Filter to Chemical Register codes only, sorted alphabetically
    filtered_codes = sorted(
        (c for c in stock_codes if c in metadata),
        key=lambda x: x.upper(),
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Chemical Register"

    NAVY = "1A2E44"   # dark navy — title row
    TEAL = "1F6B7A"   # dark teal — location/date header row
    GREEN = "2C6B33"  # CCS green — column header row
    WHITE = "FFFFFF"
    LIGHT = "F5F8FA"

    def _hdr_cell(cell, text, bg, bold=True, size=11, align="center"):
        cell.value = text
        cell.font = Font(bold=bold, size=size, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)

    def _label_cell(cell, text):
        cell.value = text
        cell.font = Font(bold=True, size=10, color="334455")
        cell.alignment = Alignment(horizontal="left", vertical="center")

    def _value_cell(cell, text):
        cell.value = text
        cell.font = Font(size=10, color="223344")
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ── Row 1: CHEMICAL REGISTER title ──────────────────────────────────────
    ws.merge_cells("A1:K1")
    _hdr_cell(ws["A1"], "CHEMICAL REGISTER", NAVY, size=16)
    ws.row_dimensions[1].height = 36

    # ── Row 2: brand sub-title ───────────────────────────────────────────────
    ws.merge_cells("A2:K2")
    ws["A2"].value = "Compliant Cleaning Supplies  |  Childcare Cleaning Supplies"
    ws["A2"].font = Font(italic=True, size=10, color="445566")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # ── Row 3: SITE-SPECIFIC label ───────────────────────────────────────────
    ws.merge_cells("A3:K3")
    ws["A3"].value = "SITE-SPECIFIC WORKPLACE CHEMICAL REGISTER"
    ws["A3"].font = Font(bold=True, size=12, color=NAVY)
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 22

    # ── Row 4: spacer ────────────────────────────────────────────────────────
    ws.row_dimensions[4].height = 8

    # ── Row 5: LOCATION / DATE / PREPARED BY labels ──────────────────────────
    ws.merge_cells("A5:B5")
    _hdr_cell(ws["A5"], "LOCATION", TEAL, size=10)
    ws.merge_cells("C5:E5")
    _hdr_cell(ws["C5"], "REGISTER CREATION DATE", TEAL, size=10)
    ws.merge_cells("F5:K5")
    _hdr_cell(ws["F5"], "PREPARED BY", TEAL, size=10)
    ws.row_dimensions[5].height = 20

    # ── Row 6: LOCATION / DATE / PREPARED BY values ──────────────────────────
    ws.merge_cells("A6:B6")
    _value_cell(ws["A6"], site_name)
    ws.merge_cells("C6:E6")
    _value_cell(ws["C6"], today)
    ws.merge_cells("F6:K6")
    _value_cell(ws["F6"], _PREPARED_BY)
    ws.row_dimensions[6].height = 20

    # ── Rows 7-10: spacers ────────────────────────────────────────────────────
    for r in range(7, 11):
        ws.row_dimensions[r].height = 8

    # ── Row 11: Column headers A–K ────────────────────────────────────────────
    headers = [
        "Product Code",
        "Chemical / Product",
        "Hazard Status",
        "UN Number",
        "Maximum Qty",
        "Risk Assessment",
        "Hazchem",
        "Class",
        "Packing Group",
        "Primary Use",
        "SDS Review Date",
    ]
    for col_idx, hdr in enumerate(headers, start=1):
        cell = ws.cell(row=11, column=col_idx)
        _hdr_cell(cell, hdr, GREEN, size=10)
    ws.row_dimensions[11].height = 22

    # ── Rows 12+: data (filtered + sorted) ────────────────────────────────────
    thin = Side(style="thin", color="D0DCE8")
    border = Border(bottom=thin)
    for row_idx, code in enumerate(filtered_codes, start=12):
        m = metadata.get(code, {})
        bg = PatternFill("solid", fgColor=LIGHT) if row_idx % 2 == 0 else None
        row_vals = [
            code,
            m.get("product_name") or "",
            m.get("hazard_classification") or "",
            m.get("un_number") or "",
            m.get("maximum_qty") or "",
            "YES" if m.get("risk_assessment_required") else "NO",
            m.get("hazchem") or "",
            m.get("chemical_class") or "",
            m.get("packing_group") or "",
            m.get("primary_use") or "",
            m.get("sds_expiry") or "",
        ]
        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border = border
            if bg:
                cell.fill = bg
        ws.row_dimensions[row_idx].height = 16

    # ── Column widths A–K ─────────────────────────────────────────────────────
    for col_letter, width in zip("ABCDEFGHIJK", [16, 32, 20, 14, 14, 18, 12, 10, 14, 22, 16]):
        ws.column_dimensions[col_letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def upload_register_to_spaces(xlsx_bytes: bytes, accno: str, date_str: str) -> str:
    """Upload per-site Chemical Register Excel to DO Spaces and return its public URL."""
    import boto3

    region = os.getenv("DO_SPACES_REGION", "syd1")
    bucket = os.getenv("DO_SPACES_BUCKET", "simplyrun-media")
    s3 = boto3.client(
        "s3",
        region_name=region,
        endpoint_url=os.getenv("DO_SPACES_ENDPOINT", f"https://{region}.digitaloceanspaces.com"),
        aws_access_key_id=os.getenv("DO_SPACES_KEY", ""),
        aws_secret_access_key=os.getenv("DO_SPACES_SECRET", ""),
    )
    key = f"ccs/registers/{date_str}/{accno}_chemical_register_{date_str}.xlsx"
    s3.put_object(
        Bucket=bucket,
        Key=key,
        Body=xlsx_bytes,
        ACL="public-read",
        ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    return f"https://{bucket}.{region}.digitaloceanspaces.com/{key}"


def parse_stock_groups(data: bytes) -> list[dict[str, Any]]:
    """Parse product-grouping.xlsx → [{primary_code, related_codes[]}]."""
    df = pd.read_excel(io.BytesIO(data), header=0, dtype=str)
    groups: list[dict[str, Any]] = []
    for _, row in df.iterrows():
        codes = [str(v).strip() for v in row if str(v).strip() and str(v).strip() != "nan"]
        if len(codes) >= 2:
            groups.append({"primary_code": codes[0], "related_codes": codes[1:]})
    return groups


# ---------------------------------------------------------------------------
# Supabase REST helpers
# ---------------------------------------------------------------------------

def _sb_url() -> str:
    return os.getenv("SUPABASE_URL", "").rstrip("/")


def _sb_headers() -> dict[str, str]:
    key = os.getenv("SUPABASE_SERVICE_ROLE_KEY", "")
    return {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Content-Type": "application/json",
        "Accept": "application/json",
    }


def _sb_post_batch(table: str, rows: list[dict], *, on_conflict: str = "merge-duplicates") -> None:
    if not rows:
        return
    url = f"{_sb_url()}/rest/v1/{table}"
    headers = {**_sb_headers(), "Prefer": f"resolution={on_conflict},return=minimal"}
    for i in range(0, len(rows), _BATCH):
        batch = rows[i : i + _BATCH]
        data = json.dumps(batch).encode()
        req = Request(url, data=data, method="POST", headers=headers)
        try:
            with urlopen(req, timeout=60):
                pass
        except HTTPError as exc:
            body = exc.read().decode()
            raise RuntimeError(f"Supabase upsert to {table} failed ({exc.code}): {body[:300]}")


def _sb_insert(table: str, rows: list[dict]) -> None:
    """Plain INSERT — no upsert Prefer header. Caller must ensure no duplicates."""
    if not rows:
        return
    url = f"{_sb_url()}/rest/v1/{table}"
    headers = {**_sb_headers(), "Prefer": "return=minimal"}
    for i in range(0, len(rows), _BATCH):
        batch = rows[i : i + _BATCH]
        data = json.dumps(batch).encode()
        req = Request(url, data=data, method="POST", headers=headers)
        try:
            with urlopen(req, timeout=60):
                pass
        except HTTPError as exc:
            body = exc.read().decode()
            raise RuntimeError(f"Supabase insert to {table} failed ({exc.code}): {body[:300]}")


def _sb_get(table: str, params: str = "") -> list[dict]:
    url = f"{_sb_url()}/rest/v1/{table}?{params}"
    req = Request(url, method="GET", headers=_sb_headers())
    try:
        with urlopen(req, timeout=30) as resp:
            return json.loads(resp.read().decode())
    except HTTPError as exc:
        raise RuntimeError(f"Supabase GET {table} failed ({exc.code}): {exc.read().decode()[:200]}")


def _sb_patch(table: str, row_filter: str, data: dict) -> None:
    url = f"{_sb_url()}/rest/v1/{table}?{row_filter}"
    req = Request(
        url, method="PATCH",
        headers={**_sb_headers(), "Content-Type": "application/json"},
        data=json.dumps(data).encode(),
    )
    try:
        with urlopen(req, timeout=15):
            pass
    except HTTPError as exc:
        raise RuntimeError(f"Supabase PATCH {table} failed ({exc.code}): {exc.read().decode()[:200]}")


def _update_last_sent_at(accno: str) -> None:
    if not accno:
        return
    try:
        _sb_patch("ccs_site_mapping", f"accno=eq.{quote(accno, safe='')}", {"last_sent_at": _now()})
    except Exception:
        pass  # non-fatal


def _sb_get_all(table: str, params: str = "") -> list[dict]:
    """Paginated fetch — returns all rows regardless of PostgREST default page size."""
    page_size = 1000
    results: list[dict] = []
    offset = 0
    while True:
        sep = "&" if params else ""
        page_params = f"{params}{sep}limit={page_size}&offset={offset}"
        batch = _sb_get(table, page_params)
        results.extend(batch)
        if len(batch) < page_size:
            break
        offset += page_size
    return results


def _sb_delete(table: str, filter_param: str) -> None:
    url = f"{_sb_url()}/rest/v1/{table}?{filter_param}"
    req = Request(url, method="DELETE", headers={**_sb_headers(), "Prefer": "return=minimal"})
    try:
        with urlopen(req, timeout=30):
            pass
    except HTTPError as exc:
        raise RuntimeError(f"Supabase DELETE {table} failed ({exc.code}): {exc.read().decode()[:200]}")


def _sb_patch(table: str, filter_param: str, data: dict) -> None:
    url = f"{_sb_url()}/rest/v1/{table}?{filter_param}"
    headers = {**_sb_headers(), "Prefer": "return=minimal"}
    req = Request(url, data=json.dumps(data).encode(), method="PATCH", headers=headers)
    try:
        with urlopen(req, timeout=30):
            pass
    except HTTPError as exc:
        raise RuntimeError(f"Supabase PATCH {table} failed ({exc.code}): {exc.read().decode()[:200]}")


# ---------------------------------------------------------------------------
# Import
# ---------------------------------------------------------------------------

def import_mapping(
    mapping_bytes: bytes | None = None,
    sds_bytes: bytes | None = None,
    risk_bytes: bytes | None = None,
    grouping_bytes: bytes | None = None,
    register_bytes: bytes | None = None,
) -> dict[str, int]:
    now = _now()

    sites: list[dict] = []
    if mapping_bytes:
        sites = parse_mapping_excel(mapping_bytes)
        # Deduplicate by accno (keep last occurrence) — duplicate accnos in one
        # batch cause PG 21000 "ON CONFLICT DO UPDATE command cannot affect row a second time"
        seen: dict[str, dict] = {}
        for s in sites:
            seen[s["accno"]] = s
        sites = list(seen.values())
        _sb_post_batch("ccs_site_mapping", [{**s, "imported_at": now} for s in sites])

    links: list[dict] = []
    if sds_bytes or risk_bytes:
        links = parse_sds_links(sds_bytes, risk_bytes)
        _sb_post_batch(
            "ccs_sds_links",
            [{**lnk, "imported_at": now} for lnk in links if lnk.get("sds_url") or lnk.get("risk_url")],
        )

    group_count = 0
    if grouping_bytes:
        groups = parse_stock_groups(grouping_bytes)
        _sb_post_batch("ccs_stock_groups", [{**g, "imported_at": now} for g in groups])
        group_count = len(groups)

    register_count = 0
    if register_bytes:
        reg_records = parse_chemical_register(register_bytes)
        # PostgREST PGRST102: all rows in a batch upsert must have identical key sets.
        # Normalize every row to the same register fields — None becomes SQL NULL,
        # which is fine since we only include register-specific columns (not sds_url/risk_url).
        _REGISTER_FIELDS = (
            "stock_code", "risk_assessment_required", "sds_expiry",
            "product_name", "hazard_classification", "primary_use",
            "signal_word", "un_number",
            "maximum_qty", "hazchem", "chemical_class", "packing_group",
        )
        # Deduplicate by stock_code — Chemical Register may list the same code
        # more than once (e.g. different size rows). Last occurrence wins.
        deduped: dict[str, dict] = {}
        for r in reg_records:
            deduped[r["stock_code"]] = {k: r.get(k) for k in _REGISTER_FIELDS}
        _sb_post_batch("ccs_sds_links", list(deduped.values()))
        register_count = len(deduped)

    # Record import event — ignore failure so it doesn't block the import response
    try:
        _sb_post_batch("ccs_import_history", [{
            "imported_at": now,
            "sites_count": len(sites),
            "sds_links_count": len(links),
            "groups_count": group_count,
            "register_count": register_count,
        }])
    except Exception:
        pass

    return {"sites": len(sites), "links": len(links), "groups": group_count, "register": register_count}


def get_import_history(limit: int = 10) -> list[dict[str, Any]]:
    """Return the most recent import events for the Data Management page."""
    rows = _sb_get("ccs_import_history", f"select=*&order=imported_at.desc&limit={limit}")
    # Mark the most recent as active, rest as superseded
    for i, row in enumerate(rows):
        row["status"] = "active" if i == 0 else "superseded"
    return rows


# ---------------------------------------------------------------------------
# Sites listing
# ---------------------------------------------------------------------------

def list_sites(search: str = "", page: int = 1, page_size: int = 50, status: str = "all", last_sent: str = "all") -> dict[str, Any]:
    from datetime import datetime, timedelta, timezone as _tz

    excl_set = {r["accno"] for r in _sb_get_all("ccs_site_exclusions", "select=accno")}
    held_set = {r["accno"] for r in _sb_get_all("ccs_site_holds", "select=accno")}

    offset = (page - 1) * page_size
    params = f"select=*&order=name.asc&limit={page_size}&offset={offset}"
    if search:
        enc = quote(search.replace("%", ""), safe="")
        params += f"&or=(name.ilike.*{enc}*,ho_name.ilike.*{enc}*)"

    # Status filter — push to PostgREST via in()/not.in()
    if status == "hold" and held_set:
        joined = ",".join(quote(a, safe="") for a in held_set)
        params += f"&accno=in.({joined})"
    elif status == "hold":
        return {"sites": [], "page": page, "page_size": page_size}
    elif status == "excluded" and excl_set:
        joined = ",".join(quote(a, safe="") for a in excl_set)
        params += f"&accno=in.({joined})"
    elif status == "excluded":
        return {"sites": [], "page": page, "page_size": page_size}
    elif status == "active":
        combined = excl_set | held_set
        if combined:
            joined = ",".join(quote(a, safe="") for a in combined)
            params += f"&accno=not.in.({joined})"

    # Last-sent filter
    now_utc = datetime.now(_tz.utc)
    if last_sent == "never":
        params += "&last_sent_at=is.null"
    elif last_sent == "this_week":
        cutoff = (now_utc - timedelta(days=7)).strftime("%Y-%m-%dT%H:%M:%SZ")
        params += f"&last_sent_at=gte.{cutoff}"
    elif last_sent == "this_month":
        cutoff = (now_utc - timedelta(days=30)).strftime("%Y-%m-%dT%H:%M:%SZ")
        params += f"&last_sent_at=gte.{cutoff}"
    elif last_sent == "older":
        cutoff = (now_utc - timedelta(days=30)).strftime("%Y-%m-%dT%H:%M:%SZ")
        params += f"&last_sent_at=lt.{cutoff}&last_sent_at=not.is.null"

    sites = _sb_get("ccs_site_mapping", params)

    for site in sites:
        site["excluded"] = site.get("accno") in excl_set
        site["held"] = site.get("accno") in held_set

    return {"sites": sites, "page": page, "page_size": page_size}


def get_stats() -> dict[str, int]:
    try:
        total = len(_sb_get_all("ccs_site_mapping", "select=accno"))
        excl = len(_sb_get_all("ccs_site_exclusions", "select=accno"))
        held = len(_sb_get_all("ccs_site_holds", "select=accno"))
        links = len(_sb_get_all("ccs_sds_links", "select=stock_code"))
        return {
            "total_sites": total,
            "excluded_sites": excl,
            "held_sites": held,
            "active_sites": total - excl - held,
            "sds_links": links,
        }
    except Exception:
        return {"total_sites": 0, "excluded_sites": 0, "held_sites": 0, "active_sites": 0, "sds_links": 0}


def get_import_status() -> dict[str, Any]:
    """Return per-table record counts and last import timestamp for the admin data panel."""
    _TABLE_PK = {
        "ccs_site_mapping": "accno",
        "ccs_sds_links": "stock_code",
        "ccs_stock_groups": "primary_code",
    }
    result: dict[str, Any] = {}
    for table, pk in _TABLE_PK.items():
        rows_all = _sb_get(table, f"select={pk},imported_at")
        last_ts = None
        for r in rows_all:
            ts = r.get("imported_at")
            if ts and (last_ts is None or ts > last_ts):
                last_ts = ts
        result[table] = {"count": len(rows_all), "last_import": last_ts}
    return result


def clear_table_data(tables: list[str]) -> dict[str, str]:
    """Delete all rows from the specified tables."""
    _ALLOWED: dict[str, str] = {
        "ccs_site_mapping": "accno",
        "ccs_sds_links": "stock_code",
        "ccs_stock_groups": "primary_code",
        "ccs_site_exclusions": "accno",
        "ccs_site_holds": "accno",
    }
    results: dict[str, str] = {}
    for table in tables:
        if table not in _ALLOWED:
            results[table] = "skipped — not allowed"
            continue
        pk = _ALLOWED[table]
        try:
            _sb_delete(table, f"{pk}=not.is.null")
            results[table] = "cleared"
        except Exception as exc:
            results[table] = f"error: {exc}"
    return results


def exclude_site(accno: str, name: str = "") -> dict[str, str]:
    _sb_post_batch("ccs_site_exclusions", [{"accno": accno, "name": name, "excluded_at": _now()}])
    return {"excluded": accno}


def include_site(accno: str) -> dict[str, str]:
    _sb_delete("ccs_site_exclusions", f"accno=eq.{quote(accno, safe='')}")
    return {"included": accno}


def hold_site(accno: str, name: str = "") -> dict[str, str]:
    _sb_post_batch("ccs_site_holds", [{"accno": accno, "name": name, "held_at": _now()}])
    return {"held": accno}


def unhold_site(accno: str) -> dict[str, str]:
    _sb_delete("ccs_site_holds", f"accno=eq.{quote(accno, safe='')}")
    return {"unheld": accno}


def preview_email(
    accno: str,
    stockcodes: list[str],
    preview_addr: str = "preview@example.com",
    *,
    public_base_url: str = "",
    tracking_secret: str = "",
) -> dict[str, Any]:
    """Compose full site email (incl. Chemical Register Excel) and return for preview without sending."""
    sites = _sb_get("ccs_site_mapping", f"select=*&accno=eq.{quote(accno, safe='')}")
    if not sites:
        raise ValueError(f"Site {accno!r} not found")
    site = sites[0]

    use_codes = stockcodes if stockcodes else (site.get("stockcodes") or [])
    if not use_codes:
        raise ValueError(f"No products specified for site {accno!r}")

    sds_map, risk_map, group_fallback, risk_required_set, register_codes = load_lookup_maps()
    docs = resolve_docs_for_site(use_codes, sds_map, risk_map, group_fallback, risk_required_set, register_codes)

    msg = compose_site_email(
        site, docs, preview_addr,
        batch_id="preview",
        public_base_url=public_base_url,
        tracking_secret=tracking_secret,
    )
    return {
        "html": msg["html"],
        "subject": msg["subject"],
        "register_url": msg.get("register_url", ""),
        "register_error": msg.get("register_error", ""),
        "docs": len(docs),
        "site_name": site.get("name", ""),
        "email": preview_addr,
    }


def send_manual(
    accno: str,
    stockcodes: list[str],
    email: str,
    dry_run: bool = False,
    *,
    public_base_url: str = "",
    tracking_secret: str = "",
) -> dict[str, Any]:
    """Send SDS/Risk email for a specific site with selected products."""
    sites = _sb_get("ccs_site_mapping", f"select=*&accno=eq.{quote(accno, safe='')}")
    if not sites:
        raise ValueError(f"Site {accno!r} not found")
    site = sites[0]

    use_codes = stockcodes if stockcodes else (site.get("stockcodes") or [])
    if not use_codes:
        raise ValueError(f"No products specified for site {accno!r}")

    sds_map, risk_map, group_fallback, risk_required_set, register_codes = load_lookup_maps()
    docs = resolve_docs_for_site(use_codes, sds_map, risk_map, group_fallback, risk_required_set, register_codes)
    if not docs:
        return {"status": "skipped", "reason": "no SDS/Risk documents found for selected products"}

    batch_id = f"manual_{accno}_{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}"
    msg = compose_site_email(
        site, docs, email,
        batch_id=batch_id,
        public_base_url=public_base_url,
        tracking_secret=tracking_secret,
    )

    if dry_run:
        return {"status": "dry_run", "site": site.get("name"), "email": email, "docs": len(docs)}

    from .distribution import _find_or_create_ghl_contact_id, _send_messages_via_ghl
    contact_id = _find_or_create_ghl_contact_id({"email": email, "name": site.get("name", "")})
    if contact_id:
        msg["contact_id"] = contact_id
    result = _send_messages_via_ghl([msg])
    _update_last_sent_at(accno)
    return {
        "status": result.get("status", "unknown"),
        "site": site.get("name"),
        "email": email,
        "docs": len(docs),
    }


# ---------------------------------------------------------------------------
# Document resolution
# ---------------------------------------------------------------------------

def load_lookup_maps() -> tuple[dict[str, str], dict[str, str], dict[str, str], set[str], set[str]]:
    """Load SDS links + stock groups into memory.

    Returns (sds_map, risk_map, group_fallback, risk_required_set, register_codes) where:
    - group_fallback maps related_code → primary_code
    - risk_required_set: codes where risk_assessment_required = True
    - register_codes: codes that are in the master Chemical Register (product_name IS NOT NULL).
      SDS-URL-only entries without a Chemical Register row are excluded intentionally.
    """
    links = _sb_get("ccs_sds_links", "select=stock_code,sds_url,risk_url,risk_assessment_required,product_name")
    sds_map = {r["stock_code"]: r["sds_url"] for r in links if r.get("sds_url")}
    risk_map = {r["stock_code"]: r["risk_url"] for r in links if r.get("risk_url")}
    risk_required_set = {r["stock_code"] for r in links if r.get("risk_assessment_required")}
    register_codes: set[str] = {r["stock_code"] for r in links if r.get("product_name")}

    groups = _sb_get("ccs_stock_groups", "select=primary_code,related_codes")
    group_fallback: dict[str, str] = {}
    for g in groups:
        primary = g.get("primary_code", "")
        for related in g.get("related_codes") or []:
            if related and related not in sds_map and related not in risk_map:
                group_fallback[related] = primary

    return sds_map, risk_map, group_fallback, risk_required_set, register_codes


def resolve_docs_for_site(
    stockcodes: list[str],
    sds_map: dict[str, str],
    risk_map: dict[str, str],
    group_fallback: dict[str, str],
    risk_required_set: set[str] | None = None,
    register_codes: set[str] | None = None,
) -> list[dict[str, str]]:
    """Resolve SDS/Risk URLs for each code.

    - Excludes codes not in the Chemical Register (register_codes) when register_codes is provided.
      A code may still be included if it resolves to a registered code via group_fallback.
    - Output is sorted alphabetically by product code.
    """
    docs: list[dict[str, str]] = []
    for code in stockcodes:
        # Filter: skip codes that are not in the Chemical Register and have no group fallback
        # to a registered code. This removes accessories/equipment from email sends.
        if register_codes is not None:
            primary_fallback = group_fallback.get(code, "")
            if code not in register_codes and primary_fallback not in register_codes:
                continue

        sds_url = sds_map.get(code, "")
        risk_url = ""
        if risk_required_set is None or code in risk_required_set:
            risk_url = risk_map.get(code, "")
        if not sds_url and not risk_url:
            primary = group_fallback.get(code, "")
            if primary:
                sds_url = sds_map.get(primary, "")
                if risk_required_set is None or code in risk_required_set:
                    risk_url = risk_map.get(primary, "")
        if sds_url or risk_url:
            docs.append({"code": code, "sds_url": sds_url, "risk_url": risk_url})

    docs.sort(key=lambda d: d["code"].upper())
    return docs


# ---------------------------------------------------------------------------
# Email composition
# ---------------------------------------------------------------------------

def compose_site_email(
    site: dict[str, Any],
    docs: list[dict[str, str]],
    email_addr: str,
    *,
    batch_id: str = "",
    public_base_url: str = "",
    tracking_secret: str = "",
    subject: str = "",
) -> dict[str, Any]:
    site_name = site.get("name", "")
    ho_name = site.get("ho_name", "") or site_name
    accno = site.get("accno", "")
    contact_id = accno or email_addr

    products_in_email = [{"code": d["code"], "name": d["code"]} for d in docs]
    documents: list[dict[str, str]] = []

    for d in docs:
        for label, url_key in [("SDS", "sds_url"), ("Risk Assessment", "risk_url")]:
            raw_url = d.get(url_key, "")
            if not raw_url:
                continue
            doc_id = f"site_{accno}_{d['code']}_{label.replace(' ', '_').lower()}"
            delivery_url = raw_url
            if public_base_url and tracking_secret:
                delivery_url = tracking_url(
                    public_base_url=public_base_url,
                    document_id=doc_id,
                    contact_id=contact_id,
                    chemical_name=d["code"],
                    redirect_url=raw_url,
                    secret=tracking_secret,
                )
            documents.append({
                "label": label,
                "product_code": d["code"],
                "chemical_name": d["code"],
                "source_url": raw_url,
                "delivery_url": delivery_url,
                "document_id": doc_id,
                "filename": raw_url.split("/")[-1],
            })

    pixel_url = ""
    if public_base_url and tracking_secret:
        pixel_url = email_open_pixel_url(
            public_base_url=public_base_url,
            email=email_addr,
            contact_id=contact_id,
            secret=tracking_secret,
            batch_id=batch_id,
        )

    logo_url = f"{public_base_url}/api/assets/ccs_logo.png" if public_base_url else ""

    html_body = _render_branded_html(
        contact_name=site_name,
        company=ho_name,
        products_in_email=products_in_email,
        documents=documents,
        tracking_pixel_url=pixel_url,
        logo_url=logo_url,
    )

    # Generate per-site Chemical Register Excel and upload to DO Spaces
    register_url = ""
    register_error = ""
    stock_codes = site.get("stockcodes") or [d["code"] for d in docs]
    if stock_codes:
        try:
            today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
            xlsx = generate_chemical_register_excel(site_name, accno, stock_codes, today)
            register_url = upload_register_to_spaces(xlsx, accno, today)
        except Exception as _exc:
            register_error = str(_exc)  # surfaced in preview; non-fatal for live send

    msg: dict[str, Any] = {
        "to": email_addr,
        "name": site_name,
        "contact_id": contact_id,
        "accno": accno,
        "subject": subject or f"Your SDS Compliance Pack — {site_name}",
        "html": html_body,
        "documents": documents,
    }
    if register_url:
        msg["attachments"] = [register_url]
        msg["register_url"] = register_url
    if register_error:
        msg["register_error"] = register_error
    return msg


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


# ---------------------------------------------------------------------------
# Beat task helpers — new product detection, expiry alerts, hold notifications
# ---------------------------------------------------------------------------

_INTERNAL_EMAIL = "ccshub@ccsessentials.com.au"
_INTERNAL_NAME = "CCS Hub Internal"


def _is_first_weekday_of_month_aest() -> bool:
    """True if today in AEST (UTC+10) is the first Mon–Fri of the calendar month."""
    import datetime as _dt
    aest = _dt.timezone(_dt.timedelta(hours=10))
    today = datetime.now(aest).date()
    if today.weekday() >= 5:
        return False
    first = today.replace(day=1)
    for delta in range((today - first).days):
        if (first + _dt.timedelta(days=delta)).weekday() < 5:
            return False
    return True


def get_new_product_queue() -> list[dict[str, Any]]:
    """Return unactioned new-product entries grouped by site (notified_at IS NULL)."""
    rows = _sb_get(
        "ccs_site_product_history",
        "select=accno,stock_code,first_seen_at&notified_at=is.null&order=first_seen_at.desc",
    )
    sites = {s["accno"]: s for s in _sb_get("ccs_site_mapping", "select=accno,name,emails")}
    by_site: dict[str, dict] = {}
    for r in rows:
        accno = r["accno"]
        if accno not in by_site:
            site = sites.get(accno, {})
            by_site[accno] = {
                "accno": accno,
                "name": site.get("name", accno),
                "emails": site.get("emails") or [],
                "products": [],
            }
        by_site[accno]["products"].append({
            "stock_code": r["stock_code"],
            "first_seen_at": r["first_seen_at"],
        })
    return list(by_site.values())


def mark_products_notified(entries: list[dict[str, Any]]) -> None:
    """Set notified_at=now for each (accno, stock_code) pair."""
    now = datetime.now(timezone.utc).isoformat()
    for e in entries:
        _sb_patch(
            "ccs_site_product_history",
            f"accno=eq.{quote(e['accno'], safe='')}&stock_code=eq.{quote(e['stock_code'], safe='')}",
            {"notified_at": now},
        )


def detect_and_record_new_products() -> dict[str, Any]:
    """Compare ccs_site_mapping stockcodes vs ccs_site_product_history.
    Insert new (accno, stock_code) pairs. Returns counts and per-site breakdown.
    On first run (empty history) seeds the table without flagging anything as new."""
    sites = _sb_get("ccs_site_mapping", "select=accno,name,stockcodes")
    # Paginated fetch — history can exceed PostgREST's default 1000-row page
    history_rows = _sb_get_all("ccs_site_product_history", "select=accno,stock_code")
    seen: set[tuple[str, str]] = {(r["accno"], r["stock_code"]) for r in history_rows}
    first_run = len(seen) == 0

    new_rows: list[dict[str, Any]] = []
    now = datetime.now(timezone.utc).isoformat()
    for site in sites:
        accno = site.get("accno", "")
        for code in (site.get("stockcodes") or []):
            if (accno, code) not in seen:
                new_rows.append({"accno": accno, "stock_code": code, "first_seen_at": now})

    if new_rows:
        # Plain INSERT — no upsert Prefer header (old PostgREST doesn't support resolution=);
        # duplicates already filtered via `seen` above so no 409 risk
        _sb_insert("ccs_site_product_history", new_rows)

    by_site: dict[str, list[str]] = {}
    for r in new_rows:
        by_site.setdefault(r["accno"], []).append(r["stock_code"])

    site_names = {s["accno"]: s.get("name", s["accno"]) for s in sites}
    return {
        "new_count": len(new_rows),
        "first_run": first_run,
        "by_site": {site_names.get(k, k): v for k, v in by_site.items()},
    }


def get_expiring_sds(days_ahead: int = 60) -> list[dict[str, Any]]:
    """Return sds_links rows with sds_expiry within days_ahead days from today."""
    import datetime as _dt
    today = datetime.now(timezone.utc).date()
    cutoff = today + _dt.timedelta(days=days_ahead)
    return _sb_get(
        "ccs_sds_links",
        f"select=stock_code,product_name,sds_expiry"
        f"&sds_expiry=gte.{today.isoformat()}"
        f"&sds_expiry=lte.{cutoff.isoformat()}"
        f"&order=sds_expiry.asc",
    )


def get_held_sites() -> list[dict[str, Any]]:
    """Return all sites currently on hold."""
    return _sb_get("ccs_site_holds", "select=accno,name,held_at&order=held_at.asc")


def send_internal_notification(subject: str, html_body: str) -> dict[str, Any]:
    """Send an internal notification email to ccshub@ccsessentials.com.au via GHL."""
    from .distribution import _find_or_create_ghl_contact_id, _send_messages_via_ghl
    contact_id = _find_or_create_ghl_contact_id({"email": _INTERNAL_EMAIL, "name": _INTERNAL_NAME}) or _INTERNAL_EMAIL
    msg: dict[str, Any] = {
        "to": _INTERNAL_EMAIL,
        "name": _INTERNAL_NAME,
        "contact_id": contact_id,
        "subject": subject,
        "html": html_body,
        "documents": [],
    }
    return _send_messages_via_ghl([msg])


def _internal_email_wrapper(title: str, body_html: str, today: str) -> str:
    return f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
  body{{font-family:Arial,sans-serif;font-size:14px;color:#222;margin:0;padding:24px;background:#f5f5f5}}
  .card{{background:#fff;border-radius:6px;padding:24px;max-width:760px;margin:0 auto;box-shadow:0 1px 4px rgba(0,0,0,.1)}}
  h2{{margin:0 0 4px;color:#2C6B33;font-size:18px}}
  .meta{{color:#888;font-size:12px;margin-bottom:20px}}
  table{{width:100%;border-collapse:collapse;margin-top:12px}}
  th{{background:#2C6B33;color:#fff;text-align:left;padding:8px 10px;font-size:13px}}
  td{{padding:7px 10px;border-bottom:1px solid #eee;font-size:13px}}
  tr:last-child td{{border-bottom:none}}
  .empty{{color:#888;font-style:italic;padding:12px 0}}
</style></head>
<body><div class="card">
<h2>{title}</h2>
<div class="meta">Generated {today} · CCS Platform</div>
{body_html}
</div></body></html>"""


def _render_new_products_email(by_site: dict[str, list[str]], today: str) -> str:
    if not by_site:
        body = '<p class="empty">No new products detected.</p>'
    else:
        rows = "".join(
            f"<tr><td>{site}</td><td>{len(codes)}</td><td style='font-size:12px'>{', '.join(codes)}</td></tr>"
            for site, codes in sorted(by_site.items())
        )
        total = sum(len(v) for v in by_site.values())
        body = (
            f"<p><strong>{total} new product–site pairs</strong> detected across {len(by_site)} site(s).</p>"
            f"<table><tr><th>Site</th><th>Count</th><th>Product Codes</th></tr>{rows}</table>"
        )
    return _internal_email_wrapper("New Products Detected", body, today)


def _render_expiry_email(products: list[dict[str, Any]], today: str) -> str:
    if not products:
        body = '<p class="empty">No products expiring within 60 days.</p>'
    else:
        rows = "".join(
            f"<tr><td>{p.get('stock_code','')}</td>"
            f"<td>{p.get('product_name') or '—'}</td>"
            f"<td>{p.get('sds_expiry','')}</td></tr>"
            for p in products
        )
        body = (
            f"<p><strong>{len(products)} product(s)</strong> with SDS expiring within 60 days.</p>"
            f"<table><tr><th>Stock Code</th><th>Product Name</th><th>Expiry Date</th></tr>{rows}</table>"
        )
    return _internal_email_wrapper("SDS Expiry Alert", body, today)


def _render_hold_list_email(sites: list[dict[str, Any]], today: str) -> str:
    if not sites:
        body = '<p class="empty">No sites currently on hold.</p>'
    else:
        rows = "".join(
            f"<tr><td>{s.get('name','')}</td><td>{s.get('accno','')}</td>"
            f"<td>{(s.get('held_at') or '')[:10]}</td></tr>"
            for s in sites
        )
        body = (
            f"<p><strong>{len(sites)} site(s)</strong> currently on hold.</p>"
            f"<table><tr><th>Site Name</th><th>Acc No</th><th>Held Since</th></tr>{rows}</table>"
        )
    return _internal_email_wrapper("Weekly Hold List", body, today)
